import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
import os
import datetime
import logging
import logging.handlers
from django.shortcuts import render
from django.http import HttpResponse
from app2.forms import ReportForm
import json
from report import views
from config.settings import BASE_DIR
import shutil
# import re




class Searching_qiita():
    def qiita_search(self, keys, pages):
        #loggerを設定
        logger=logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        #Handlerを設定
        ima=datetime.datetime.now()
        filename='Qiita_log {0} {1}h-{2}m-{3}s.log'.format(ima.strftime('%Y-%m-%d'),ima.hour, ima.minute, ima.second)
        h1=logging.handlers.RotatingFileHandler(filename, maxBytes=100000, backupCount=10)
        h1.setLevel(logging.DEBUG)
        #HandlerのFormatを設定
        formatter=logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        h1.setFormatter(formatter)
        #loggerにHandlerを設定
        logger.addHandler(h1)
        try:
    
            now=datetime.datetime.now()
            logger.debug('キーボード入力/情報取得開始')    
            # keys=input('検索したい文字を入力してください。')
            # pages=input('ページ数を数字で入力してください。')
            results=requests.get('https://qiita.com/search?q={}'.format(keys))
            soup=BeautifulSoup(results.text, 'lxml')
            logger.debug('キーボード入力/情報取得終了')
    
            logger.debug('情報のリスト化/抽出開始')
            list_title=[]
            list_setsumei=[]
            list_date=[]
            list_title=soup.find_all(True, {'class':'searchResult_itemTitle'})#見出し、URLを含むclassを指定
            list_setsumei=soup.find_all(True, {'class':'searchResult_snippet'})
            # list_date=soup.find_all(True, {"class":"searchResult_header"})
            # kensuu_data=soup.find_all(True, {'class':'badge'})　検索結果の全記事数を表示。今回は、検索結果の記事数のみでいいので一旦コメント化して無効に
            # kensuu=int(re.sub("\\D", "", str(kensuu_data)))
            logger.debug('情報の抽出完了')
            
            qiita_titles=[] #検索結果の、タイトル、URL、記事冒頭のそれぞれを入れるリストを用意。
            qiita_urls=[]
            qiita_setsumei=[]
            qiita_results=[]
            qiita_date=[]

            for x in range(int(pages)):
                logger.debug('ページの出力開始')
                if len(list_title)==0:
                    logger.debug('結果出力開始2')
                    qiita_titles=[]
                    qiita_urls=[]
                    qiita_setsumei=[]
                    qiita_date=[]
                    kensuu=[]
                    file_title_q=""
                    qiita_results.insert(0, qiita_titles)
                    qiita_results.insert(1, qiita_urls)
                    qiita_results.insert(2, qiita_setsumei)
                    qiita_results.insert(3, file_title_q)
                    qiita_results.insert(4, kensuu)
                    qiita_results.insert(5, qiita_date)
                    logger.debug('結果出力終了2')
                    return qiita_results

                for i in range(len(list_title)):#上記リストの要素数の分だけ繰り返す
                    url=list_title[int(i)].find('a')#find_allで抽出した要素から、さらに<a>タグのみを抽出
                    link=url.get('href')#<a>タグ内から、urlが格納されているhrefを取得
                    title=url.get_text()#<a>タグの中のテキスト部分＝見出しを取得
                    setsumei=list_setsumei[int(i)].get_text()
                    qiita_urls.append('https://qiita.com'+link)#<a>タグのhrefには、urlの一部のみ格納されていたので、https://...を追加する
                    qiita_titles.append(title)
                    qiita_setsumei.append(setsumei)
                    results2=requests.get('{}'.format(qiita_urls[int(i)]))
                    soup2=BeautifulSoup(results2.text, 'lxml')
                    hidsuke=soup2.find("time")
                    hidsuke_2=hidsuke.get("datetime")
                    hidsuke_3=hidsuke_2.replace('T', " ")
                    hidsuke_4=hidsuke_3.replace('Z', "")
                    # hidsuke=list_date[int(i)].get_text()
                    # hidsuke_2=hidsuke.lstrip("が").rstrip("に投稿")
                    qiita_date.append(hidsuke_4)
                    

                logger.debug('ページの出力完了')
                #ここから下は、次ページへの移動
                logger.debug('次ページの出力開始')
                list_title.clear()
                list_setsumei.clear()
                list_date.clear()
                next_page=soup.find(True, {'class':'js-next-page-link'})#次ページへ移動するURLを含むclassを指定
                if not next_page==None:
                    url_next_page=next_page.get('href')
                    results=requests.get('https://qiita.com/'+url_next_page)#次ページへ移動するURLに、http://qiita.com/を追加
                    soup=BeautifulSoup(results.text, 'lxml')
                    list_title=soup.find_all(True, {'class':'searchResult_itemTitle'})#このプログラム冒頭と同様に、見出し、URLを含むclassを指定
                    list_setsumei=soup.find_all(True, {'class':'searchResult_snippet'})
                    list_date=soup.find_all(True, {"class":"searchResult_header"})
                    print(url_next_page)
                    now=datetime.datetime.now()
                    logger.debug('次ページの情報取得完了')
                else:
                    logger.debug('全ページの情報取得処理完了')
                    break
            file_title_q='media/Qiita検索結果({0}) {1} {2}h{3}m{4}s.xlsx'.format(keys,now.strftime('%Y-%m-%d'),now.hour, now.minute, now.second)
            for z in range(len(qiita_titles)):
                if not os.path.isfile(file_title_q):#保存先エクセルファイルがない場合、新規で作成する。
                    wb_qiita=openpyxl.Workbook()
                    sheet=wb_qiita.active
                    sheet.title='Qiita検索結果'#シート名を設定
                    wb_qiita.save(file_title_q)
                ws=wb_qiita['Qiita検索結果']
                cell_title='A'+str(z+2)#エクセルのAセルに、処理を繰り返すたびに上から値を入れていく。中身は、見出し。+2となっているのは、後から1行目にインデックスを追加するため。
                cell_link='B'+str(z+2)#こちらは、BセルにURLを入れていく。
                ws[cell_link]=qiita_urls[int(z)]
                ws[cell_link].hyperlink='https://qiita.com'+qiita_urls[int(z)]#ハイパーリンク化
                ws[cell_link].font=openpyxl.styles.fonts.Font(color='0000FF')#リンクらしく青色に
                ws[cell_title]=qiita_titles[int(z)]
            
                cell_setsumei='C'+str(z+2)
                ws[cell_setsumei]=qiita_setsumei[int(z)]
                cell_date="D"+str(z+2)
                ws[cell_date]=qiita_date[int(z)]
                    
                ws['A1']='タイトル'# エクセル1行目に各インデックスを付ける
                ws['C1']='記事冒頭'
                ws['B1']='URL'
                ws['D1']='記事作成日'
                wb_qiita.save(file_title_q)
            logger.debug('全処理完了')
            # with open('qiita_title_data.json', 'w') as ft, open('qiita_links_data.json', 'w') as fl, open('qiita_setsumei_data.json', 'w') as fs:
            # #ここの２つの処理は、title, linkそれぞれを新たに入れたリストを外部ファイルとして書き込み、外部ファイルを保存している処理。
            #     json.dump(qiita_title_info, ft, ensure_ascii=False)
            #     json.dump(qiita_links_info, fl, ensure_ascii=False)
            #     json.dump(qiita_setsumei_info, fs, ensure_ascii=False) #ここでおそらく文字コードが一致していない？？
            logger.debug('結果出力開始1')
            kensuu=len(qiita_titles)
            qiita_results.insert(0, qiita_titles)
            qiita_results.insert(1, qiita_urls)
            qiita_results.insert(2, qiita_setsumei)
            qiita_results.insert(3, file_title_q)
            qiita_results.insert(4, kensuu)
            qiita_results.insert(5, qiita_date)
            logger.debug('結果出力終了1')
            # logger.debug('結果出力開始2')
            # os.mkdir('./media/')
            # shutil.move(file_title, './media/')
            # logger.debug('結果出力終了2')
            return qiita_results
            
        except UnboundLocalError as ULE:
            logger.debug('UnboudLocalError発生')
            print('UnboundLocalError発生')
            print(ULE)
        except ValueError as VE:
            logger.debug('ValueError発生')
            print('ValueError発生')
            print(VE)
        except Exception as e:
            logger.debug('一部の処理でエラーが発生しています。')
            print('[Error]',e)
            print('一部の処理でエラーが発生しています。')
        
            
# if __name__=='__main__':
#     Searching_qiita()


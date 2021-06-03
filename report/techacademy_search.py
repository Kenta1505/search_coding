import requests
from bs4 import BeautifulSoup
import openpyxl
# from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import datetime
import logging
import logging.handlers
from django.shortcuts import render
from django.http import HttpResponse
from app2.forms import ReportForm
import json
from report import views
# import codecs



class Searching_ta():
    def ta_search(self, keys, pages):
        #loggerを設定
        logger=logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        #Handlerを設定
        ima=datetime.datetime.now()
        filename='TechAcademy_log {0} {1}h-{2}m-{3}s.log'.format(ima.strftime('%Y-%m-%d'),ima.hour, ima.minute, ima.second)
        h1=logging.handlers.RotatingFileHandler(filename, maxBytes=100000, backupCount=10)
        h1.setLevel(logging.DEBUG)
        #HandlerのFormatを設定
        formatter=logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        h1.setFormatter(formatter)
        #loggerにHandlerを設定
        logger.addHandler(h1)
        try:
            logger.debug('キーボード入力/情報取得開始')    
            # keys=input('検索したい文字を入力してください。')
            # pages=input('ページ数を数字で入力してください')
            results=requests.get('https://techacademy.jp/magazine/?s={}'.format(keys))
            soup=BeautifulSoup(results.text, 'lxml')
            logger.debug('キーボード入力/情報取得終了')

            logger.debug('情報のリスト化/抽出開始')
            
            ta_titles=[] #検索結果の、タイトル、URL、記事冒頭のそれぞれを入れるリストを用意。
            ta_urls=[]
            ta_setsumei=[]
            ta_results=[]
            ta_date=[]

            list_title=[]
            list_setsumei=[]
            list_title=soup.find_all(True, {'class':'entry-title'})#見出し、URLを含むclassを指定
            list_setsumei=soup.find_all(True, {'class':'entry-excerpt'})
            now=datetime.datetime.now()
            logger.debug('情報の抽出完了')
            
            # logger.debug('JSONファイル作成開始')
            # if not os.path.isfile('ta_title_data.json'):
            #     #検索結果を入れるリストを外部ファイル化する処理。外部ファイルがなければ、新しく作る。
            #     myfile_t=open('ta_title_data.json', 'w')
            #     json.dump(ta_titles, myfile_t)
            #     myfile_t.close
            # logger.debug('JSONファイル for Title作成完了')
            
            # if not os.path.isfile('ta_links_data.json'):
            #     #検索結果を入れるリストを外部ファイル化する処理。外部ファイルがなければ、新しく作る。
            #     myfile_l=open('ta_links_data.json', 'w')
            #     json.dump(ta_urls, myfile_l)
            #     myfile_l.close
            # logger.debug('JSONファイル for links作成完了')
            
            # if not os.path.isfile('ta_setsumei_data.json'):
            #     #検索結果の記事冒頭部分を入れる外部ファイルを作成する処理。
            #     myfile_s=open('ta_setsumei_data.json','w')
            #     json.dump(ta_setsumei, myfile_s)
            #     myfile_s.close
            # logger.debug('JSONファイル for 記事冒頭作成完了')
            
            # logger.debug('Open JSON files')
            
            # myfile_t=open('ta_title_data.json', 'r')
            # myfile_l=open('ta_links_data.json', 'r')
            # myfile_s=open('ta_setsumei_data.json', 'r')
            
            # logger.debug('JSONファイル読み込み開始')
            
            # ta_title_info=json.load(myfile_t)
            # ta_links_info=json.load(myfile_l)
            # ta_setsumei_info=json.load(myfile_s)
            
            # logger.debug('JSONファイル読み込み完了')
            
            # logger.debug('JSONファイル内リストクリーン開始')
            
            # ta_title_info.clear() #外部ファイルがすでにある場合、前回の検索結果がリストに残っている可能性があるため、一度clear()を行う。
            # ta_links_info.clear() #外部ファイルがすでにある場合、前回の検索結果がリストに残っている可能性があるため、一度clear()を行う。
            # ta_setsumei_info.clear() #外部ファイルがすでにある場合、前回の検索結果がリストに残っている可能性があるため、一度clear()を行う。
            
            # logger.debug('JSONファイル準備完了')

            # if keys in list_title:
            for x in range(int(pages)):
                for i in range(len(list_title)-10):#上記リストの要素数の分だけ繰り返す
                    url=list_title[int(i)].find('a')#find_allで抽出した要素から、さらに<a>タグのみを抽出
                    link=url.get('href')#<a>タグ内から、urlが格納されているhrefを取得
                    title=url.get_text()#<a>タグの中のテキスト部分＝見出しを取得
                    ta_urls.append(link)
                    ta_titles.append(title)
                    results2=requests.get('{}'.format(ta_urls[int(i)]))
                    soup2=BeautifulSoup(results2.text, 'lxml')
                    hidsuke=soup2.find(True, {"class":"date"})
                    hidsuke2=hidsuke.get_text()
                    ta_date.append(hidsuke2)

                
                
                for y in range(len(list_setsumei)):
                    logger.debug('ページ見出しの出力開始')
                    #TechAcademyのサイトだと、検索ワードのページ以外に特集記事もスクレイピングされるため、
                    #タイトルと見出しの数が合わないので分けた。
                    setsumei=list_setsumei[int(y)].get_text()
                    ta_setsumei.append(setsumei) #ここの文字コードがおかしい？？あるいは、strip()でもう少し整理が必要？？

                    logger.debug('ページ見出しの出力完了')
                
                logger.debug('次ページの出力開始')
                list_title.clear()
                list_setsumei.clear()
                next_page=soup.find(True, {'class':'nav-links'})#見出し、URLを含むclassを指定
                if not next_page==None:
                    a_next_page=next_page.find_all(True, {'class':'next page-numbers'})
                    url_next_page=a_next_page[0].get('href')
                    results=requests.get(url_next_page)
                    soup=BeautifulSoup(results.text, 'lxml')
                    list_title=soup.find_all(True, {'class':'entry-title'})#見出し、URLを含むclassを指定
                    list_setsumei=soup.find_all(True, {'class':'entry-excerpt'})
                    print(url_next_page)
                    now=datetime.datetime.now()
                    logger.debug('次ページの出力完了')
                else:
                    logger.debug('全ページの情報取得処理完了')
                    break
                
            file_title_ta='media/TechAcademy検索結果({0}) {1} {2}h{3}m{4}s.xlsx'.format(keys,now.strftime('%Y-%m-%d'),now.hour, now.minute, now.second)
            for z in range(len(ta_titles)):   
                cell_title='A'+str(z+2)
                #エクセルのAセルに、処理を繰り返すたびに上から値を入れていく。中身は、見出し
                #+2となっているのは、後から1行目にインデックスを追加するため。
                logger.debug('ファイル作成開始')
                if not os.path.isfile(file_title_ta):#保存先エクセルファイルがない場合、新規で作成する。
                    wb_ta=openpyxl.Workbook()
                    sheet=wb_ta.active
                    sheet.title='TechAcademy検索結果'#シート名を設定
                    wb_ta.save(file_title_ta)
                    ws=wb_ta['TechAcademy検索結果']
                logger.debug('ファイル作成完了')
                    #複数ファイル生成の場合の、処理について要検討。スクレイピング後、ダウンロードできるファイルは一つになっている。ダウンロードボタンを複数設置するか、検索結果を一つのファイルにまとめるか、複数のファイルをZIP形式などでまとめてダウンロードできるようにするか。
                logger.debug('セル入力開始 -title-')
                ws[cell_title]=ta_titles[int(z)]
                logger.debug('title入力完了')
                
                logger.debug('url入力開始')
                cell_link='B'+str(z+2)#こちらは、BセルにURLを入れていく。
                ws[cell_link]=ta_urls[int(z)]
                ws[cell_link].hyperlink=link#ハイパーリンク化
                ws[cell_link].font=openpyxl.styles.fonts.Font(color='0000FF')#リンクらしく青色に
                logger.debug('url入力完了')
                
                logger.debug('A1, B1セル入力開始')
                ws['A1']='タイトル'
                ws['B1']='URL'
                logger.debug('A1, B1完了')
                wb_ta.save(file_title_ta)
                
            for y in range(len(ta_setsumei)):
                logger.debug('説明入力開始')
                cell_setsumei='C'+str(y+2)
                ws[cell_setsumei]=ta_setsumei[int(y)]
                logger.debug('説明入力完了')
                wb_ta.save(file_title_ta)
                logger.debug('ファイル作成完了')
                
                #この下は、上記A,Bのセルの列幅を自動調整する内容を作成しようとしたが、未完成。主要部分ではないので、一旦後回し。
        
                # ws1=wb_qiita.worksheets[0]
                # for col in ws1.columns:
                    #     max_length=0
                    #     column = col[0].columns
                
                #     for cell in col:
                    #         if len(str(cell.value)) > max_length:
                        #             max_length = len(str(cell.value))
                        # adjusted_width=(max_length + 2) * 1.2
                        # ws1.column_dimensions[get_column_letter(column)].width=adjusted_width
                ws['C1']='記事冒頭'
                wb_ta.save(file_title_ta)
                logger.debug('ページの出力完了')
    
            logger.debug('全処理完了')
            # else:
            #     ta_titles=""
            #     ta_urls=""
            #     ta_setsumei=""
            #     file_title_ta=""
            # with open('ta_title_data.json', 'w') as ft, open('ta_links_data.json', 'w') as fl, open('ta_setsumei_data.json', 'w') as fs:
            # #ここの２つの処理は、title, linkそれぞれを新たに入れたリストを外部ファイルとして書き込み、外部ファイルを保存している処理。
            #     json.dump(ta_title_info, ft, ensure_ascii=False)
            #     json.dump(ta_links_info, fl, ensure_ascii=False)
            #     json.dump(ta_setsumei_info, fs, ensure_ascii=False) #ここでおそらく文字コードが一致していない？？
            kensuu=len(ta_titles)
            ta_results.insert(0, ta_titles)
            ta_results.insert(1, ta_urls)
            ta_results.insert(2, ta_setsumei)
            ta_results.insert(3, file_title_ta)
            ta_results.insert(4, kensuu)
            ta_results.insert(5, ta_date)
            return ta_results
                

        except UnboundLocalError as ULE:
            logger.debug('UnboudLocalError発生')
            print('UnboundLocalError発生')
            print(ULE)
        except ValueError as VE:
            logger.debug('ValueError発生')
            print('ValueError発生')
            print(VE)
        except Exception as e:
            logger.debug('一部の処理でエラーが発生しています。')
            print('[Error]',e)
            print('一部の処理でエラーが発生しています。')

    
        
# if __name__=='__main__':
#     Searching_ta()


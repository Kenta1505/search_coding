import requests
from bs4 import BeautifulSoup
import openpyxl
# from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import datetime
import logging
import logging.handlers
from django.shortcuts import render
from django.http import HttpResponse
from app2.forms import ReportForm
import json
from report import views



class Searching_samurai():
    def samurai_search(self, keys, pages):
        #loggerを設定
        logger=logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        #Handlerを設定
        ima=datetime.datetime.now()
        filename='Samurai_log {0} {1}h-{2}m-{3}s.log'.format(ima.strftime('%Y-%m-%d'),ima.hour, ima.minute, ima.second)
        h1=logging.handlers.RotatingFileHandler(filename, maxBytes=100000, backupCount=10)
        h1.setLevel(logging.DEBUG)
        #HandlerのFormatを設定
        formatter=logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        h1.setFormatter(formatter)
        #loggerにHandlerを設定
        logger.addHandler(h1)
        try:
            now=datetime.datetime.now()
            logger.debug('検索文字列入力/情報取得開始')    
            # keys=request.POST['search']
            # pages=request.POST['pages']
            results=requests.get('https://www.sejuku.net/blog/?s={}'.format(keys))
            soup=BeautifulSoup(results.text, 'lxml')
            logger.debug('キーボード入力/情報取得完了')    
            # wb_qiita_soup=openpyxl.Workbook() -->この部分は、soupで取り出した内容を確認・保存したいときに使う。
            # sheet=wb_qiita_soup.active
            # sheet.title='Samurai_soup'#シート名を設定
            # wb_qiita_soup.save('Samurai_soup({0}) {1} {2}h{3}m{4}s.xlsx'.format(keys,now.strftime('%Y-%m-%d'),now.hour, now.minute, now.second))
            # ws_s=wb_qiita_soup['Samurai_soup']
            # ws_s['A1']=soup.prettify()
            # wb_qiita_soup.save('Samurai_soup({0}) {1} {2}h{3}m{4}s.xlsx'.format(keys,now.strftime('%Y-%m-%d'),now.hour, now.minute, now.second))
            logger.debug('情報のリスト化/抽出開始')    
            # print(soup.prettify())
            list_title=[]
            # list_category=[]
            list_title=soup.find_all(True, {'class':'entry-title'})#見出し、URLを含むclassを指定
            # list_category=soup.find_all(True, {'class':'meta-category'})
            samurai_titles=[]
            samurai_urls=[]
            samurai_results=[]
            samurai_date=[]
            logger.debug('情報の抽出完了')
            
            # logger.debug('JSONファイル作成開始')
            # if not os.path.isfile('sm_title_data.json'):
            #     #検索結果を入れるリストを外部ファイル化する処理。外部ファイルがなければ、新しく作る。
            #     myfile_t=open('sm_title_data.json', 'w')
            #     json.dump(samurai_titles, myfile_t)
            #     myfile_t.close
            # logger.debug('JSONファイル for Title作成完了')
            
            # if not os.path.isfile('sm_links_data.json'):
            #     #検索結果を入れるリストを外部ファイル化する処理。外部ファイルがなければ、新しく作る。
            #     myfile_l=open('sm_links_data.json', 'w')
            #     json.dump(samurai_urls, myfile_l)
            #     myfile_l.close
            # logger.debug('JSONファイル for links作成完了')
            # logger.debug('Open JSON files')
            
            # myfile_t=open('sm_title_data.json', 'r')
            # myfile_l=open('sm_links_data.json', 'r')
            
            # logger.debug('JSONファイル読み込み開始')
            
            # sm_title_info=json.load(myfile_t)
            # sm_links_info=json.load(myfile_l)
            
            # logger.debug('JSONファイル読み込み完了')
            
            # logger.debug('JSONファイル内リストクリーン開始')
            
            # sm_title_info.clear() #外部ファイルがすでにある場合、前回の検索結果がリストに残っている可能性があるため、一度clear()を行う。
            # sm_links_info.clear() #外部ファイルがすでにある場合、前回の検索結果がリストに残っている可能性があるため、一度clear()を行う。
            
            # logger.debug('JSONファイル準備完了')
            
            for x in range(int(pages)):
                logger.debug('ページの出力開始')
                if len(list_title)==0:
                    logger.debug('結果出力開始-データなし-')
                    samurai_titles=[]
                    samurai_urls=[]
                    kensuu=[]
                    samurai_date=[]
                    file_title_sm=""
                    samurai_results.insert(0, samurai_titles)
                    samurai_results.insert(1, samurai_urls)
                    samurai_results.insert(2, file_title_sm)
                    samurai_results.insert(3, kensuu)
                    samurai_results.insert(4, samurai_date)
                    logger.debug('結果出力終了-データなし-')
                    return samurai_results

                for i in range(len(list_title)):#上記リストの要素数の分だけ繰り返す
                    url=list_title[int(i)].find('a')#find_allで抽出した要素から、さらに<a>タグのみを抽出
                    link=url.get('href')#<a>タグ内から、urlが格納されているhrefを取得
                    title=url.get_text()#<a>タグの中のテキスト部分＝見出しを取得
                    # 記事のcategoryを取得しようとしたが、うまくいかず。とりあえず、後回しに。
                    # category=list_category[int(i)].find_all('a')
                    # cate=category.get_text()
                    # c=cate.get_text()
                    samurai_titles.append(title) #スクレイピング結果をリストに入れる。
                    samurai_urls.append(link) #スクレイピング結果をリストに入れる。
                    results2=requests.get('{}'.format(samurai_urls[int(i)]))
                    soup2=BeautifulSoup(results2.text, 'lxml')
                    hidsuke=soup2.find(True, {"class":"meta-author-modified-date"})
                    hidsuke2=hidsuke.get_text()
                    samurai_date.append(hidsuke2)


                    # print(category)　記事のカテゴリ取得に失敗。一旦後回し
                logger.debug('ページの出力完了')
                
                #ここから下は、次ページへの移動
                logger.debug('次ページの出力開始')
                list_title.clear()
                next_page=soup.find(True, {'class':'next page-numbers'})#次ページへ移動するURLを含むclassを指定
                if not next_page==None:
                    url_next_page=next_page.get('href')
                    results=requests.get(url_next_page)
                    soup=BeautifulSoup(results.text, 'lxml')
                    list_title=soup.find_all(True, {'class':'entry-title'})#プログラム冒頭と同じく、見出し、URLを含むclassを指定
                    print(url_next_page)
                    now=datetime.datetime.now()
                    logger.debug('次ページの情報取得完了')
                else:
                    logger.debug('全ページの情報取得処理完了')
                    break
            logger.debug('ファイル出力開始')
            file_title_sm='media/Samurai検索結果({0}) {1} {2}h{3}m{4}s.xlsx'.format(keys,now.strftime('%Y-%m-%d'),now.hour, now.minute, now.second)
            for z in range(len(samurai_titles)):
                logger.debug('ファイル作成開始')
                if not os.path.isfile(file_title_sm):#保存先エクセルファイルがない場合、新規で作成する。
                    wb_samurai=openpyxl.Workbook()
                    sheet=wb_samurai.active
                    sheet.title='Samurai検索結果'#シート名を設定
                    wb_samurai.save(file_title_sm)
                logger.debug('ファイル作成完了')
                logger.debug('セル入力開始 -title-')
                ws=wb_samurai['Samurai検索結果']
                cell_title='A'+str(z+2)#エクセルのAセルに、処理を繰り返すたびに上から値を入れていく。中身は、見出し。+2となっているのは、後から1行目にインデックスを追加するため。
                ws[cell_title]=samurai_titles[int(z)]
                logger.debug('title入力完了')
                # cell_category='B'+str(i+1)
                # ws[cell_category]=cate
                logger.debug('セル入力開始-url-')
                cell_link='B'+str(z+2)#こちらは、BセルにURLを入れていく。
                ws[cell_link]=samurai_urls[int(z)]
                ws[cell_link].hyperlink=samurai_urls[int(z)]#ハイパーリンク化
                ws[cell_link].font=openpyxl.styles.fonts.Font(color='0000FF')#リンクらしく青色に
                logger.debug('url入力完了')
                #この下は、上記A,Bのセルの列幅を自動調整する内容を作成しようとしたが、未完成。主要部分ではないので、一旦後回し。
    
                # ws1=wb_qiita.worksheets[0]
                # for col in ws1.columns:
                #     max_length=0
                #     column = col[0].columns
                    
                #     for cell in col:
                #         if len(str(cell.value)) > max_length:
                #             max_length = len(str(cell.value))
                # adjusted_width=(max_length + 2) * 1.2
                # ws1.column_dimensions[get_column_letter(column)].width=adjusted_width
                logger.debug('A1,B1入力開始')
                ws['A1']='タイトル'#エクセル1行目にインデックスを付ける
                ws['B1']='URL'
                logger.debug('A1, B1入力完了')
                wb_samurai.save(file_title_sm)
                logger.debug('エクセルファイル作成完了')
            logger.debug('全処理完了')
            # print(samurai_titles)
            # print(samurai_urls)
            # print(title_info, links_info)
            # with open('sm_title_data.json', 'w') as ft, open('sm_links_data.json', 'w') as fl:
            # #ここの２つの処理は、title, linkそれぞれを新たに入れたリストを外部ファイルとして書き込み、外部ファイルを保存している処理。
            #     json.dump(sm_title_info, ft, ensure_ascii=False)
            #     json.dump(sm_links_info, fl, ensure_ascii=False)
            if not samurai_titles==[]:
                kensuu=len(samurai_titles)
            else:
                kensuu=[]
            samurai_results.insert(0, samurai_titles)
            samurai_results.insert(1, samurai_urls)
            samurai_results.insert(2, file_title_sm)
            samurai_results.insert(3, kensuu)
            samurai_results.insert(4, samurai_date)
            return samurai_results

        except UnboundLocalError as ULE:
            logger.debug('UnboudLocalError発生')
            print('UnboundLocalError発生')
            print(ULE)
        except ValueError as VE:
            logger.debug('ValueError発生')
            print('ValueError発生')
            print(VE)
        except Exception as e:
            logger.debug('一部の処理でエラーが発生しています。')
            print('[Error]',e)
            print('一部の処理でエラーが発生しています。')

# if __name__=='__main__':
#     Searching_samurai()


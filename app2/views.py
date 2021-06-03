from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.template import Context, loader
from app2.forms import ReportForm, PageForm
from report.samurai_search import Searching_samurai
from report.techacademy_search import Searching_ta
from report.qiita_search import Searching_qiita
import openpyxl
from openpyxl.styles import Font
import os
import datetime
from openpyxl.writer.excel import save_virtual_workbook
import logging
import logging.handlers
from django.views import generic
# from .models import UploadFile
from django.http import FileResponse

# def index(request):
#     template=loader.get_template('index.html')
#     context={}
#     return HttpResponse(template.render(context, request))

logger=logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
#Handlerを設定
ima=datetime.datetime.now()
filename='LOG {0} {1}h-{2}m-{3}s.txt'.format(ima.strftime('%Y-%m-%d'),ima.hour, ima.minute, ima.second)
h1=logging.handlers.RotatingFileHandler(filename, maxBytes=100000, backupCount=10)
h1.setLevel(logging.DEBUG)
#HandlerのFormatを設定
formatter=logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
h1.setFormatter(formatter)
#loggerにHandlerを設定
logger.addHandler(h1)



def searching(request):
    f=ReportForm()
    p=PageForm()
    context_form={'form1':f,'form2':p}
    if request.method=='POST':
        if 'search_button' in request.POST:
            # context_results={}
            kensaku_words=str(request.POST['search'])
            kensaku_pages=int(request.POST['pages'])
            f=ReportForm({'search':kensaku_words})
            p=PageForm({'pages':kensaku_pages})
            context_form={'form1':f, "form2":p}
            checkbox_target_sm=request.POST.get("checkbox_sites_sm", None)
            checkbox_target_ta=request.POST.get("checkbox_sites_ta", None)
            checkbox_target_qiita=request.POST.get("checkbox_sites_qiita", None)
            
            if type(checkbox_target_sm)==str and checkbox_target_sm in 'samurai':
                S=Searching_samurai()
                s_results=S.samurai_search(kensaku_words, kensaku_pages)
                sm_data=[]
                logger.debug("侍 スクレイピング処理完了。出力開始")
                if not s_results[0]==[] or not s_results[1]==[]:
                    logger.debug("侍 記事が見つかった場合の処理開始")
                    for (x_s, y_s, d_s) in zip(s_results[0], s_results[1], s_results[4]):
                        c_s=[x_s, y_s, d_s]
                        sm_data.append(c_s)
                    context_samurai={
                        'context_samurai':sm_data
                    }
                    excel_samurai={
                        'excel_samurai':s_results[2]
                    }
                    kensuu_samurai={
                        'kensuu_samurai':s_results[3]
                    }
                    context_form.update(context_samurai)
                    context_form.update(excel_samurai)
                    context_form.update(kensuu_samurai)
                else:
                    logger.debug("侍 記事が見つからない場合の処理開始")
                    context_samurai={
                        'dummy':'samurai'
                    }
                    context_form.update(context_samurai)
            else:
                pass
            
            if type(checkbox_target_ta)==str and checkbox_target_ta in 'ta':
                T=Searching_ta()
                t_results=T.ta_search(kensaku_words, kensaku_pages)
                ta_data=[]
                logger.debug("TA スクレイピング処理完了。出力開始")
                if not t_results[0]==[] or not t_results[1]==[] or not t_results[2]==[] or not type(t_results[0])==None or not type(t_results[1])==None or not type(t_results[2])==None:
                    logger.debug("TA 記事が見つかった場合の処理開始")
                    if t_results[4]==0:
                        logger.debug("TA 記事が見つからない場合の処理開始 パターン１")
                        context_ta={
                            'dummy':'ta'
                        }
                        ta_data.clear
                        # t_excel=t_results[3].clear()
                        # excel_ta={
                        #     'excel_ta':t_excel
                        # }
                        # print(t_excel)
                        context_form.update(context_ta)
                        # context_form.update(excel_ta)
                    else:
                        for (x_t, y_t, z_t, d_t) in zip(t_results[0], t_results[1], t_results[2], t_results[5]):
                            c_t=[x_t, y_t, z_t, d_t]
                            ta_data.append(c_t)
                        context_ta={
                            'context_ta':ta_data
                        }
                        excel_ta={
                            'excel_ta':t_results[3]
                        }
                        kensuu_ta={
                            'kensuu_ta':t_results[4]
                        }
                        context_form.update(context_ta)
                        context_form.update(excel_ta)
                        context_form.update(kensuu_ta)
                else:
                    logger.debug("TA 記事が見つからない場合の処理開始　パターン２")
                    context_ta={
                        'dummy':'ta'
                    }
                    t_excel=t_results[3].clear()
                    excel_ta={
                        'excel_ta':t_excel
                    }
                    print(t_excel)
                    context_form.update(context_ta)
                    context_form.update(excel_ta)
            else:
                pass
            
            if type(checkbox_target_qiita)==str and checkbox_target_qiita in 'qiita':
                Q=Searching_qiita()
                q_results=Q.qiita_search(kensaku_words, kensaku_pages)
                qiita_data=[]
                logger.debug("Qiita スクレイピング処理完了。出力開始")
                if not q_results[0]==[] or not q_results[1]==[] or not q_results[2]==[]:
                    logger.debug("Qiita 記事が見つかった場合の処理開始")
                    for (x_q, y_q, z_q, w_q) in zip(q_results[0], q_results[1], q_results[2], q_results[5]):
                        c_q=[x_q, y_q, z_q, w_q]
                        qiita_data.append(c_q)
                    context_qiita={
                        'context_qiita':qiita_data
                    }
                    excel_qiita={
                        'excel_qiita':q_results[3]
                    }
                    kensuu_qiita={
                        'kensuu_qiita':q_results[4]
                    }
                    context_form.update(context_qiita)
                    context_form.update(excel_qiita)
                    context_form.update(kensuu_qiita)
                    
                else:
                    logger.debug("Qiita 記事が見つからない場合の処理開始")
                    context_qiita={
                        'dummy':"qiita"
                    }
                    context_form.update(context_qiita)
            else:
                pass
            
            logger.debug("全体 contextの出力完了。templateに飛ばします。")
            context=(context_form)
            return render(request, 'top.html', context)

    else:
        f=ReportForm({
            'title':'レポートタイトル',
            'body':'Hello. Django! Form',
            }
        )
        context=(context_form)
        return render(request, 'top.html', context)
    context=(context_form)
    return render(request, 'top.html', context)
    



# def file_download(request):
#     logging.debug('開始1')
#     # if request.method=="POST":
#         # if 'search_button' in request.POST:
#     logging.debug('開始2')
#     # file_title=q_results[3]
#     # if not os.path.isfile(file_title):#保存先エクセルファイルがない場合、新規で作成する。
#     #     wb_qiita=openpyxl.Workbook()
#     #     sheet=wb_qiita.active
#     #     sheet.title='Qiita検索結果'#シート名を設定
#     #     wb_qiita.save(file_title)
#     # ws=wb_qiita['Qiita検索結果']
#     # logging.debug('終了2')
#     # logging.debug('開始3')
#     # for i in range(len(qiita_data[0])):
#     #     cell_title='A'+str(i+2)#エクセルのAセルに、処理を繰り返すたびに上から値を入れていく。中身は、見出し。+2となっているのは、後から1行目にインデックスを追加するため。
#     #     cell_link='B'+str(i+2)#こちらは、BセルにURLを入れていく。
#     #     ws[cell_link]=qiita_data[1][int(i)]
#     #     ws[cell_link].hyperlink=qiita_data[1][int(i)]#ハイパーリンク化
#     #     ws[cell_link].font=openpyxl.styles.fonts.Font(color='0000FF')#リンクらしく青色に
#     #     ws[cell_title]=qiita_data[0][int(i)]
    
#     #     cell_setsumei='C'+str(i+2)
#     #     ws[cell_setsumei]=qiita_data[2][int(i)]
        
#     #     ws['A1']='タイトル'# エクセル1行目に各インデックスを付ける
#     #     ws['C1']='記事冒頭'
#     #     ws['B1']='URL'
#     #     wb_qiita.save(file_title)
#     #     logging.debug('終了3')
#     # logging.debug('開始4')
#     wb_qiita=openpyxl.load_workbook(file_title)
#     response = HttpResponse(content=save_virtual_workbook(wb_qiita), content_type='application/vnd.ms-excel')
#     response['Content-Disposition'] = 'attachment; filename={}'.format(file_title)
#     wb_qiita.save(response)
#     # context=({"excel_file":"/testsite2/{title}".format(file_title)})
#     logging.debug('終了4')
#     logging.debug('終了1')
#     return response

# class UploadList(generic.ListView):
#     model=UploadFile

# def download(request, pk):
#     upload_file=get_object_or_404(UploadFile, pk=pk)
#     file=upload_file.file
#     return FileResponse(file)